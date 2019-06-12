import urllib
import openpyxl
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import load_workbook
#soup = BeautifulSoup()



theurl = "http://www.ntz.de/service/anzeigenservice/kleinanzeigen/marktplatz/"

thepage = urllib.request.urlopen(theurl)
soup = BeautifulSoup(thepage,"html.parser")

print(soup.title.text)
""""
for link in soup.findAll('a'):
    #print(link.get('href'))
    print(link.text)
"""
#print(soup.find('div',{"class":"tx-ntzkleinanz-pi1"}).find('article',{"class":"teaser"}).find('p').text)

wb = load_workbook(filename = 'test.xlsx')

dest_filename = 'test.xlsx'

ws1 = wb.active
ws1.title = "WebData"

i = 0
liste = []
for anzeigen in soup.findAll('article',{"class":"teaser"}):
    #print(anzeigen)
    print(anzeigen.find('p').text)
    i = i + 1
    liste.append(str(i) + ". " + str(anzeigen.find('p').text))

i=1
for Anzeige in liste:
    ws1.cell(row=i, column=1).value = Anzeige
    #print(liste[i])
    i=i+1


wb.save(filename = dest_filename)




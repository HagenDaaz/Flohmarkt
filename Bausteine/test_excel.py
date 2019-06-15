from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import urllib.request
from bs4 import BeautifulSoup

wb = load_workbook(filename = 'test.xlsx')

dest_filename = 'test.xlsx'

ws1 = wb.active
ws1.title = "WebData"

#ws2 = wb.create_sheet(title="Hagen")

theurl = "http://www.ntz.de/service/anzeigenservice/kleinanzeigen/marktplatz/"

thepage = urllib.request.urlopen(theurl)
soup = BeautifulSoup(thepage,"html.parser")

print(soup.title.text)
""""
for link in soup.findAll('a'):
    #print(link.get('href'))
    print(link.text)
"""
#print(soup.find('div',{"class":"tx-ntzkleinanz-pi1"}).find('article',{"class":"teaser"}).find('p').text)

liste = []
for i, anzeigen in enumerate(soup.findAll('article',{"class":"teaser"})):
    number = i + 1
    text = str(number) + ". " + str(anzeigen.find('p').text)
    #print(number)
    liste.append(text)
    ws1.cell(row=i+1, column=1).value = text

wb.save(filename = dest_filename)

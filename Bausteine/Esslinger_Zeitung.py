import urllib
import openpyxl
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import os

#soup = BeautifulSoup()
i = 0

main_site = "https://www.esslinger-zeitung.de/"
markt_site = "".join([main_site,"/anzeigensuche_costart,"])

theurl1 = "".join([markt_site, "1_pageid,anzeigensuche_rubrik,1172.html"])
theurl2 = "".join([markt_site, "11_pageid,anzeigensuche_rubrik,1172.html"])
theurl3 = "".join([markt_site, "21_pageid,anzeigensuche_rubrik,1172.html"])
theurl4 = "".join([markt_site, "31_pageid,anzeigensuche_rubrik,1172.html"])
theurl5 = "".join([markt_site, "41_pageid,anzeigensuche_rubrik,1172.html"])
theurl6 = "".join([markt_site, "51_pageid,anzeigensuche_rubrik,1172.html"])
theurl7 = "".join([markt_site, "61_pageid,anzeigensuche_rubrik,1172.html"])

urls = ([theurl1, theurl2, theurl3, theurl4, theurl5, theurl6, theurl7])

def harvest_url(urls=urls):
    liste = []
    papernameliste = []
    for eachurl in urls:
        # print("Seite",Seitennr)
        thepage = urllib.request.urlopen(eachurl)
        soup = BeautifulSoup(thepage, "html.parser")

        for anzeigen in soup.findAll('section', {"class": "nfy-c-adv nfy-c-adv-search-teaser cf"}):
            # print("Anzeige ",Anzeigennr)
            print(anzeigen.find('p').text)
            liste.append(anzeigen.find('p').text)  # .split('\n') #''.join(

    return liste


# wb = load_workbook(filename = 'test.xlsx')
#
# dest_filename = 'test.xlsx'
#
# ws1 = wb.active
# ws1.title = "WebData"
#
# i=0
#
# for Anzeige in liste:
#     ws1.cell(row=i+1, column=1).value = Anzeige
#     print(liste[i])
#     i=i+1
#
# wb.save(filename=dest_filename)
# """"
# for link in soup.findAll('a'):
#     #print(link.get('href'))
#     print(link.text)
# """
#print(soup.find('div',{"class":"tx-ntzkleinanz-pi1"}).find('article',{"class":"teaser"}).find('p').text)



    #liste.append(str(i) + ". " + str(anzeigen.find('p').text))

#print(liste[0])

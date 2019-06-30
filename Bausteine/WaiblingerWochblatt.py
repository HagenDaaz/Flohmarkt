import urllib.request
from bs4 import BeautifulSoup
from PIL import Image
import pytesseract
import io
from openpyxl import load_workbook

main_site = "https://www.mein-wochenblatt.de"

markt_site = "".join([main_site,"/index.php?WBID=&kat="])

theurl1 = "".join([markt_site, "64&r1=9000&r2=9100"])
theurl2 = "".join([markt_site, "64&r1=9000&r2=9800"])

urls = ([theurl1, theurl2])
def harvest_url(urls=urls):
    pageno=1
    liste=[]


    for eachurl in urls:
        #print("Seite ", pageno)
        pageno = pageno + 1
        thepage = urllib.request.urlopen(eachurl)
        soup = BeautifulSoup(thepage,"html.parser")
        #print(soup.title.text)
        liste.append(soup.title.text)

        def text_from_img_url(img_url):
            """
            return ocr output from image
            """
            with urllib.request.urlopen(img_url) as url:

                f = io.BytesIO(url.read())
                # deutsches Sprachpaket ausgewählt
                return pytesseract.image_to_string(Image.open(f), lang='deu')


        for anzeige in soup.findAll('div', {"class": "anzeigenkasten"}):
            img_url=anzeige.find('img').get('src')
            #print(img_url)
            liste.append(str(img_url))
            img_text = text_from_img_url(img_url)
            liste.append(str(img_text))
            #print(img_text)

    return liste

# wb = load_workbook(filename = 'test.xlsx')
#
# dest_filename = 'test.xlsx'
#
# ws1 = wb.active
# ws1.title = "WebData"
#
# i=0
# for Anzeige in liste:
#     ws1.cell(row=i+1, column=1).value = Anzeige
#     print(liste[i])
#     i=i+1
#
# wb.save(filename=dest_filename)
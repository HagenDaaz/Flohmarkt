from PIL import Image
import pytesseract
import os
from PIL import Image
import pytesseract
from openpyxl import load_workbook


format = ".JPG"
#format = ".png"

liste = []
for file in os.listdir():
    jpg=file.endswith(format)
    if jpg == True:
        pfad = os.path.abspath(".") + "\\"+ file
        liste.append(pfad)
        print(os.path.abspath(".") + "\\"+ file)
        im = Image.open(file)
        text = pytesseract.image_to_string(im, lang='deu')
        liste.append(text)
        print(text)

wb = load_workbook(filename = 'test.xlsx')

dest_filename = 'test.xlsx'

ws1 = wb.active
ws1.title = "WebData"

i=0
for Anzeige in liste:
    if i % 2:
        ws1.cell(row=i+1, column=1).value = Anzeige
        print(liste[i])
        i=i+1
    else:
        ws1.cell(row=i + 1, column=1).hyperlink = Anzeige
        print(liste[i])
        i = i + 1

wb.save(filename=dest_filename)
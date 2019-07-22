import urllib.request
from bs4 import BeautifulSoup
from PIL import Image
import pytesseract
import io
from openpyxl import load_workbook

main_site = "https://www.mein-mittwochmarkt.de"

markt_site = "".join([main_site,"/Marktplatz/"])

theurl1 = "".join([markt_site, "dies-und-das.html?sl=60400"])
theurl2 = "".join([markt_site, "dies-und-das-Part2.html?sl=60400&nPart=2"])
theurl3 = "".join([markt_site, "dies-und-das-Part3.html?sl=60400&nPart=3"])
theurl4 = "".join([markt_site, "dies-und-das-Part4.html?sl=60400&nPart=4"])
theurl5 = "".join([markt_site, "dies-und-das-Part5.html?sl=60400&nPart=5"])
theurl6 = "".join([markt_site, "dies-und-das-Part6.html?sl=60400&nPart=6"])
theurl7 = "".join([markt_site, "dies-und-das-Part7.html?sl=60400&nPart=7"])
theurl8 = "".join([markt_site, "dies-und-das-Part8.html?sl=60400&nPart=8"])
theurl9 = "".join([markt_site, "dies-und-das-Part9.html?sl=60400&nPart=9"])
theurl10 = "".join([markt_site, "dies-und-das-Part10.html?sl=60400&nPart=10"])
theurl11 = "".join([markt_site, "dies-und-das.html?sl=70000"])
theurl12 = "".join([markt_site, "dies-und-das-Part2.html?sl=70000&nPart=2"])
theurl13 = "".join([markt_site, "dies-und-das-Part3.html?sl=70000&nPart=3"])
theurl14 = "".join([markt_site, "dies-und-das-Part4.html?sl=70000&nPart=4"])

urls = ([theurl1, theurl2, theurl3, theurl4, theurl5, theurl6, theurl7, theurl8, theurl9, theurl10, theurl11, theurl12, theurl13, theurl14])

def text_from_img_url(img_url):
    """
    return ocr output from image
    """
    with urllib.request.urlopen(img_url) as url:

        f = io.BytesIO(url.read())
        # deutsches Sprachpaket ausgewählt
        return pytesseract.image_to_string(Image.open(f), lang='deu')

def harvest_url(urls=urls):
    pageno = 1
    liste = []

    for eachurl in urls:
        print("Seite ", pageno)
        pageno = pageno + 1
        thepage = urllib.request.urlopen(eachurl)
        soup = BeautifulSoup(thepage, "html.parser")
        print(soup.title.text)
        liste.append(soup.title.text)

        for anzeige in soup.findAll('div', {"class": "MarketSearchCtrl_ResultList_Image"}):
            anzeige_url = anzeige.find('a').get('href')
            if anzeige_url:
                # alle anzeigen, wenn anzeige_url
                this_url = "".join([main_site, anzeige_url])
                sub_soup = BeautifulSoup(urllib.request.urlopen(this_url), "html.parser")

                for item in sub_soup.findAll('div', {"class": "lightBoxDiv"}):
                    img_url = "".join([markt_site, item.find('a').get('href')])
                    print(img_url)
                    liste.append(str(img_url))
                    img_text = text_from_img_url(img_url)
                    liste.append(str(img_text))

                    if len(img_text.split('\n')) <= 10:
                            print(img_text, )
    return liste

# wb = load_workbook(filename = 'test.xlsx')
#
# dest_filename = 'test.xlsx'
#
# ws1 = wb.active
# ws1.title = "WebData"
#
# i=0
# for Anzeige in liste:
#     ws1.cell(row=i+1, column=1).value = Anzeige
#     print(liste[i])
#     i=i+1
#
# wb.save(filename=dest_filename)

#theurl2 = "https://www.mein-mittwochmarkt.de/Marktplatz/Motiv-m128823.html?from=29.05.2019&to=04.06.2019&sort=tblMotif.dtmWebBegin+desc&sl=60400&cid=70300&sl=60400&branch="
#for image in soup.findAll('div',{"class":"lightBoxDiv"}):
#        print(image)
#for link in soup.findAll('a'):
#    variable = link.get('href')
#    print(variable)



#https://www.mein-mittwochmarkt.de/Marktplatz/Motif/Korb-Rattan-TruhenAuswahl-Korbhaus-Tue-Kilchberg-Bahnhofs--v3-w1024-h-m128814.jpg


#for img in soup.findAll('img'):
#print(img.get('src'))

#thepage = urllib.request
#soupdata=BeautifulSoup(thepage, "html.paser")
#soup = make_soup




#resource = urllib.request.urlopen("https://www.mein-mittwochmarkt.de/Marktplatz/Motif/Verkaufe-Schmuck-und-Gemaelde--v3-w1024-h-m128823.jpg")
#output = open("file01.jpg","wb")
#output.write(resource.read())
#output.close()

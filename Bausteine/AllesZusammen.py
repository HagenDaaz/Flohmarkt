from openpyxl import load_workbook
import urllib.request
from bs4 import BeautifulSoup

#Definition der zu findenden Begriffe
begriff1= "Musikinstrumente"
begriff2= "Wössner"
begriff3= "Alles" \
          ""
begriffsliste = [begriff1,begriff2,begriff3]

#Öffnen der Exceldatei
wb = load_workbook(filename = 'test1.xlsx')
#Speichername
dest_filename = 'test1.xlsx'
#
ws1 = wb.active
ws1.title = "Anzeigen"
#neuen Reiter anlegen
#ws2 = wb.create_sheet(title="Hagen")

i = 0
#Durchsuchen der Homepage
theurl = "http://www.ntz.de/service/anzeigenservice/kleinanzeigen/marktplatz/"

thepage = urllib.request.urlopen(theurl)
soup = BeautifulSoup(thepage,"html.parser")

#Zeitungstitel ausgeben
print(soup.title.text)

#print(soup.find('div',{"class":"tx-ntzkleinanz-pi1"}).find('article',{"class":"teaser"}).find('p').text)

#Befüllung der anzeigenliste
anzeigenliste = []
for anzeigen in soup.findAll('article',{"class":"teaser"}):
    #print(i)
    #print(anzeigen.find('p').text)
    i = i + 1
    anzeigenliste.append(str(i) + ". " + str(anzeigen.find('p').text))
j=1
#Überprüfung ob Begriff in Anzeigenliste enthalten ist
for anzeige in anzeigenliste:
    for begriff in begriffsliste:
        if begriff in anzeige:
            print("gefunden!")
            ws1.cell(row=j, column=1).value = anzeige
            j=j+1
            print(anzeige)
        else:
            print("nicht gefunden!")


wb.save(filename = dest_filename)
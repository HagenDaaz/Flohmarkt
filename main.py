import os, sys
import urllib
import openpyxl
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pprint import pprint
import re
from datetime import date


_basePath = os.path.realpath(os.path.dirname(__file__))
sys.path.append(_basePath)

from Bausteine.xlsxReader import XlsxReader
from Bausteine.Esslinger_Zeitung import harvest_url as esz_liste
from Bausteine.Schw_Tagblatt import harvest_url as schw_tag_liste
from Bausteine.WaiblingerWochblatt import harvest_url as waib_liste
from Bausteine.NuertingerZeitung import harvest_url as nuert_liste
from Bausteine.PforzheimerZeitung import harvest_url as pz_liste
from Bausteine.Suedwest_Presse import harvest_url as swp_liste


alleAnzeigen=[]
papername=[]

papername1=[]
liste1 = esz_liste()
esz_name = "Esslinger Zeitung"
for Anzeige in liste1:
    papername1.append(esz_name)

papername2=[]
liste2 = schw_tag_liste()
schw_tag_name = "Schwäbisches Tagblatt"
for Anzeige in liste2:
    papername2.append(schw_tag_name)

papername3=[]
liste3 = waib_liste()
waib_name = "Waiblinger Wochenblatt"
for Anzeige in liste3:
    papername3.append(waib_name)

papername4=[]
liste4 = nuert_liste()
nuert_name = "Nürtinger Zeitung"
for Anzeige in liste4:
    papername4.append(nuert_name)

papername5=[]
liste5 = pz_liste()
pz_name = "Pforzheimer Zeitung"
for Anzeige in liste5:
    papername5.append(pz_name)

papername6=[]
liste6 = swp_liste()
swp_name = "Südwestpresse"
for Anzeige in liste6:
    papername6.append(swp_name)

alleAnzeigen = liste1 + liste2 + liste3 + liste4 + liste5 + liste6
papername = papername1 + papername2 + papername3 + papername4 + papername5 + papername6
# alleAnzeigen = liste1 + liste2 + liste3 + liste4
# alleAnzeigen = liste5
# papername = papername5

i=0
for Anzeige in alleAnzeigen:
    #print(alleAnzeigen[i])
    i=i+1

wb = load_workbook(filename = 'test.xlsx')

dest_filename = 'test.xlsx'

ws1 = wb.active
ws1.title = "WebData"

for i, anzeige in enumerate(alleAnzeigen):

    # suchtext = re.search('st', anzeige, flags=0)
    # if suchtext is not None:
    #     #print(dir(suchtext))
    #     #exit()
    #     ws1.cell(row=i+1, column=1).value = suchtext.group(0)
    #
    # ws1.cell(row=i+1, column=4).value = anzeige
    ws1.cell(row=i + 1, column=1).value = date.today()
    ws1.cell(row=i + 1, column=3).value = anzeige

for i, name in enumerate(papername):
    ws1.cell(row=i + 1, column=2).value = name


    #print(alleAnzeigen[i])

wb.save(filename=dest_filename)
